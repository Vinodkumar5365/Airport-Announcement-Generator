from os import name
import tkinter as tk
from tkinter import *
from tkinter import ttk,messagebox
from PIL import Image,ImageTk
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import tkinter.ttk
import pandas as pd
from pydub import AudioSegment
from gtts import gTTS
import random
import mysql.connector
db_connection = mysql.connector.connect(
    host = "localhost",
    user = "root",
    password = ""
)
db_cursor = db_connection.cursor()
class LoginApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Login")
        self.geometry("1368x768+0+0") #initialization of size of window
        
        #backgrond image
        self.bg=ImageTk.PhotoImage(file="/home/vinodkumar/Vinod/Project/4th Sem/Project/Comparison Plane.jpg")
        bg=Label(self,image=self.bg).place(x=0,y=0)

        #register frame inside the window
        frame1=Frame(self,bg="white")
        frame1.place(x=30,y=30,width=600,height=680)

        #contents
        title=Label(frame1,text="Login",font=("times new roman",40,"bold","underline"),bg="white",fg="black").place(x=50,y=50)
        email=Label(frame1,text="Enter E-mail :",font=("times new rom an",20,"bold"),bg="white",fg="black").place(x=50,y=190)
        password=Label(frame1,text="Enter Password :",font=("times new roman",20,"bold"),bg="white",fg="black").place(x=50,y=380)

        #creates the entry fields
        self.txt_mail=Entry(frame1,font=("times new roman",15),bg="lightgrey",bd="1")
        self.txt_mail.place(x=50,y=260,width=500)

        def togglep():
            if self.txt_pass.cget('show') == '':
                self.txt_pass.config(show='*')
            else:
                self.txt_pass.config(show='')

        self.txt_pass=Entry(frame1,font=("times new roman",15),bg="lightgrey",bd="1")
        self.txt_pass.place(x=50,y=450,width=460)
        self.toggle_btn=ImageTk.PhotoImage(file="/home/vinodkumar/Vinod/Project/4th Sem/Project/eye.png")
        toggle_btn = Button(frame1,text=' ',bg="white",image=self.toggle_btn,bd="2",cursor="hand2", command=togglep).place(x=520,y=450,width=30,height=30)

        #creates the buttons
        btn=Button(frame1,text="LOGIN",font=("times new roman",20,"bold"),bg="#13114f",fg="white",bd="4",cursor="hand2",command=self.login).place(x=50,y=570,width=225)
        btn=Button(frame1,text="REGISTER",font=("times new roman",20,"bold"),bg="#13114f",fg="white",bd="4",cursor="hand2",command=self.open_registration_window).place(x=325,y=570,width=225)

    def open_registration_window(self):#opens the register window
        self.withdraw()
        window = RegisterWindow(self)
        window.grab_set()

    def show(self):
        """"""
        self.update()
        self.deiconify()

    def login(self):
        if self.txt_mail.get()=="" or self.txt_pass.get()=="":#checking the fields are empty or not
            messagebox.showerror("Error","All fields are Mandatory",parent=self)
        else :
            try :
                db_cursor.execute("use airportstaffs")#uses the database
                db_cursor.execute("select * from registers where email=%s and password=%s",(self.txt_mail.get(),self.txt_pass.get()))
                row=db_cursor.fetchone()
                if row==None:
                    messagebox.showerror("Error","Invalid Email and Password",parent=self)#checks the fields are valid are not
                else :
                    self.withdraw()
                    self.txt_mail.delete(0,END)
                    self.txt_pass.delete(0,END)
                    window = Login_Success_Window(self)#enters the main generator page
                    window.grab_set()
            except Exception as es :
                messagebox.showerror("Error",f"Error Due to : {str(es)}",parent=self)

    def exit(self):
        MsgBox = messagebox.askquestion('Exit Application', 'Are you sure you want to exit the application',icon='warning')
        if MsgBox == 'yes':
            self.destroy()

class Login_Success_Window(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.original_frame = parent
        self.title("Bangalore Airport")
        self.geometry("1368x768+0+0") #initialization of size of window

        self.bg=ImageTk.PhotoImage(file="/home/vinodkumar/Vinod/Project/4th Sem/Project/Color Planes.jpg")
        bg=Label(self,image=self.bg).place(x=-320,y=0)

        frame1=Frame(self,bg="white")
        frame1.place(x=456,y=0,width=912,height=760)

        title=Label(frame1,text="Welcome to Airport",font=("times new roman",30,"bold","underline"),bg="white",fg="black").place(x=303,y=10)

        #connection to the xl sheet
        wb = Workbook()
        wb = load_workbook("Flight.xlsx")
        ws = wb.active

        #connecting to each columns
        column_a = ws["A"]
        column_b = ws["B"]
        column_c = ws["C"]
        column_d = ws["D"]

        #getting the data on each cell
        def get_a():
            list=""
            for cell in column_a:
                list=f"{list + str(cell.value)}\n"
            label_1.config(text=list)
            list=""
            for cell in column_b:
                list=f"{list + str(cell.value)}\n"
            label_2.config(text=list)
            list=""
            for cell in column_c:
                list=f"{list + str(cell.value)}\n"
            label_3.config(text=list)
            list=""
            for cell in column_d:
                list=f"{list + str(cell.value)}\n"
            label_4.config(text=list)

        ba = Button(frame1 , text="Show Flight Details", font=("times new roman",18,"bold"),bg="#009933",fg="white",bd="4",cursor="hand2",command=get_a)
        ba.place(x=140,y=80,width=250)

        #labels to seperate each columns
        label_1=Label(frame1,text="",bg="white")
        label_1.place(x=100,y=150)
        label_2=Label(frame1,text="",bg="white")
        label_2.place(x=320,y=150)
        label_3=Label(frame1,text="",bg="white")
        label_3.place(x=540,y=150)
        label_4=Label(frame1,text="",bg="white")
        label_4.place(x=752,y=150)

        btn=Button(frame1,text="START",font=("times new roman",18,"bold"),bg="#009933",fg="white",bd="4",cursor="hand2",command=self.startgen).place(x=522,y=80,width=250)
        btn=Button(frame1,text="LOGOUT",font=("times new roman",18,"bold"),bg="#009933",fg="white",bd="4",cursor="hand2",command=self.logout).place(x=331,y=660,width=250)

    def logout(self):#enters into login page
        MsgBox = messagebox.askquestion('Exit Application', 'Are you sure you want to Logout',icon='warning')
        if MsgBox == 'yes':
            self.destroy()
            self.original_frame.show()

    def startgen(self):#it will convert the text gathered from the xl sheet
        def textToSpeech(text, filename):
            mytext = str(text)
            language = 'en'
            myobj = gTTS(text=mytext, lang=language, slow=False)
            myobj.save(filename)
    
        # This function returns pydubs audio segment
        def mergeAudios(audios):
            combined = AudioSegment.empty()
            for audio in audios:
                combined += AudioSegment.from_mp3(audio)
            return combined

        def generateSkeleton():
            audio = AudioSegment.from_mp3('flight.mp3')

            # 1 - Generate  intro music
            start = 3610
            finish = 6860
            audioProcessed = audio[start:finish]
            audioProcessed.export("1_en.mp3", format="mp3")

            # 2 - Generate  May i have your attention please
            start = 10760
            finish = 12600
            audioProcessed = audio[start:finish]
            audioProcessed.export("2_en.mp3", format="mp3")

            # 3 - Generate passengers for flight 
            start = 12650
            finish = 14000
            audioProcessed = audio[start:finish]
            audioProcessed.export("3_en.mp3", format="mp3")

            # 4 flight no

            # 5 - Generate to 
            start = 15310
            finish = 15900
            audioProcessed = audio[start:finish]
            audioProcessed.export("5_en.mp3", format="mp3")

            # 6 to city

            # 7 - Generate  is arriving shortly , please proceed to gate
            start = 16450
            finish = 19350
            audioProcessed = audio[start:finish]
            audioProcessed.export("7_en.mp3", format="mp3")

            # 8 gate no
            
            # 9 - Generate   thank you 
            start = 20130
            finish = 20900
            audioProcessed = audio[start:finish]
            audioProcessed.export("9_en.mp3", format="mp3")

            # 10 - Generate  b/w music
            start = 3610
            finish = 6860
            audioProcessed = audio[start:finish]
            audioProcessed.export("10_en.mp3", format="mp3")

            #11 - Generate Kripya Dhyan dijiye
            start = 23200
            finish = 24900
            audioProcessed = audio[start:finish]
            audioProcessed.export("11_en.mp3", format="mp3")
            
            #12 to city
            
            #13 - generate ko jaane wale gadi sankhya
            start = 25470
            finish = 27250
            audioProcessed = audio[start:finish]
            audioProcessed.export("13_en.mp3", format="mp3")
            
            #14 flight no
            
            #15 - generate Khushi samay mein platform sankhya
            start = 29100
            finish = 31400
            audioProcessed = audio[start:finish]
            audioProcessed.export("15_en.mp3", format="mp3")
            
            #16 gate no
            
            #17 generate per a rahi hai
            start = 31750
            finish = 32830
            audioProcessed = audio[start:finish]
            audioProcessed.export("17_en.mp3", format="mp3")
            
            #18 thank you
            start = 33650
            finish = 34700
            audioProcessed = audio[start:finish]
            audioProcessed.export("18_en.mp3", format="mp3")

            # 19 - Generate  b/w music
            start = 3610
            finish = 6860
            audioProcessed = audio[start:finish]
            audioProcessed.export("19_en.mp3", format="mp3")
            
        def generateAnnouncement(filename):
            df = pd.read_excel(filename)
            for index, item in df.iterrows():
                # 4 - Generate flight_no
                textToSpeech(item['flightno'], '4_en.mp3')
                # 6 - Generate to-city
                textToSpeech(item['tocity'], '6_en.mp3')
                # 8 - Generate gate_no
                textToSpeech(item['gate'], '8_en.mp3')
                # 12 - generate to city
                textToSpeech(item['tocity'], '12_en.mp3')
                # 14 - Generate flight_no
                textToSpeech(item['flightno'], '14_en.mp3')
                # 16 - Generate gate_no
                textToSpeech(item['gate'], '16_en.mp3')

                audios = [f"{i}_en.mp3" for i in range(1,20)]
                announcement = mergeAudios(audios)
                announcement.export(f"announcement_{item['flightno']}{index+1}.mp3", format="mp3")

        if __name__ == "__main__":
            generateSkeleton()
            generateAnnouncement("Flight.xlsx")

class RegisterWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.original_frame = parent
        self.title("Registeration Window")
        self.geometry("1368x768+0+0") #initialization of size of window

        self.bg=ImageTk.PhotoImage(file="/home/vinodkumar/Vinod/Project/4th Sem/Project/Moon Plane.jpg")
        bg=Label(self,image=self.bg).place(x=0,y=0)

        #register frame inside the window
        frame2=Frame(self,bg="white")
        frame2.place(x=50,y=280,width=1268,height=420)

        #contents
        title=Label(frame2,text="Register",font=("times new roman",30,"bold","underline"),bg="white",fg="black").place(x=40,y=30)

        #Labels of the text boxes and entry fields
        eid=Label(frame2,text="Enter EID :",font=("times new roman",18,"bold"),bg="white",fg="black").place(x=40,y=110)
        self.txt_eid=Entry(frame2,font=("times new roman",15),bg="lightgrey",bd="1")
        self.txt_eid.place(x=40,y=150,width=360)

        # def get_id():
        #     for x in range(1):
        #         self.txt_eid==(random.randint(1, 101))
        # getid_btn = Button(frame2,text='>',bg="#2f0340",fg="white",bd="2",cursor="hand2", command=get_id).place(x=370,y=150,width=30,height=30)

        name=Label(frame2,text="Enter Name :",font=("times new roman",18,"bold"),bg="white",fg="black").place(x=454,y=110)
        self.txt_name=Entry(frame2,font=("times new roman",15),bg="lightgrey",bd="1")
        self.txt_name.place(x=454,y=150,width=360)

        mobile=Label(frame2,text="Enter Mobile No. :",font=("times new roman",18,"bold"),bg="white",fg="black").place(x=868,y=110)
        self.txt_mobile=Entry(frame2,font=("times new roman",15),bg="lightgrey",bd="1")
        self.txt_mobile.place(x=868,y=150,width=360)

        email=Label(frame2,text="Enter E-mail :",font=("times new roman",18,"bold"),bg="white",fg="black").place(x=40,y=230)
        self.txt_email=Entry(frame2,font=("times new roman",15),bg="lightgrey",bd="1")  
        self.txt_email.place(x=40,y=270,width=774)

        password=Label(frame2,text="Enter Password :",font=("times new roman",18,"bold"),bg="white",fg="black").place(x=868,y=230)
        def togglepa():
            if self.txt_password.cget('show') == '':
                self.txt_password.config(show='*')
            else:
                self.txt_password.config(show='')
        
        self.txt_password=Entry(frame2,font=("times new roman",15),bg="lightgrey",bd="1")
        self.txt_password.place(x=868,y=270,width=320)
        self.toggle_btn=ImageTk.PhotoImage(file="/home/vinodkumar/Vinod/Project/4th Sem/Project/eye.png")
        toggle_btn = Button(frame2,text=' ',bg="white",image=self.toggle_btn,bd="2",cursor="hand2", command=togglepa).place(x=1198,y=270,width=30,height=30)

        #buttons
        bt1=Button(frame2,text="LOGIN",font=("times new roman",18,"bold"),bg="#2f0340",fg="white",bd="4",cursor="hand2",command=self.onClose).place(x=124,y=340,width=250)
        bt2=Button(frame2,text="CLEAR",font=("times new roman",18,"bold"),bg="#2f0340",fg="white",bd="4",cursor="hand2",command=self.clear_form).place(x=509,y=340,width=250)
        bt3=Button(frame2,text="REGISTER",font=("times new roman",18,"bold"),bg="#2f0340",fg="white",bd="4",cursor="hand2",command=self.register_data).place(x=894,y=340,width=250)

    #clears the contents on click
    def clear_form(self):
        self.txt_eid.delete(0,tkinter.END)
        self.txt_name.delete(0,tkinter.END)
        self.txt_email.delete(0,tkinter.END)
        self.txt_mobile.delete(0,tkinter.END)
        self.txt_password.delete(0,tkinter.END)

    #data insertion on database
    def register_data(self):
        db_cursor.execute("CREATE DATABASE IF NOT EXISTS airportstaffs")
        db_cursor.execute("use airportstaffs")#uses database
        db_cursor.execute("Create table if not exists registers(eid VARCHAR(10) NOT NULL  PRIMARY KEY,name VARCHAR(30),email VARCHAR(30),mobile VARCHAR(10),password VARCHAR(15))")
        db_connection.commit()#enters the data

        if self.txt_eid.get()=="" or self.txt_name.get()=="" or self.txt_email.get()=="" or self.txt_mobile.get()=="" or self.txt_password.get()=="":
            messagebox.showerror("Error","All fields are Mandatory",parent=self)
        else:
            if self.txt_name.get()<='6' or self.txt_email.get()<'10' or self.txt_mobile.get()!='10':
                messagebox.showerror("Error","Enter Correct details in Respective fields",parent=self)
            else:
                db_cursor.execute("use airportstaffs")
                query ="INSERT INTO registers(eid,name,email,mobile,password) VALUES ('%s','%s','%s','%s','%s')" %(self.txt_eid.get(),self.txt_name.get(),self.txt_email.get(),self.txt_mobile.get(),self.txt_password.get())
                db_connection.commit()
                try:
                    db_cursor.execute(query)
                    messagebox.showinfo('Information', "Data inserted Successfully")
                    db_connection.commit()
                    #clearing the contents of the register page after data inserted
                    self.txt_eid.delete(0,tkinter.END)
                    self.txt_name.delete(0,tkinter.END)
                    self.txt_email.delete(0,tkinter.END)
                    self.txt_mobile.delete(0,tkinter.END)
                    self.txt_password.delete(0,tkinter.END)
                except:
                    messagebox.showinfo('Information', "Data insertion failed!!!")
                    db_connection.rollback()
                    db_connection.close()

    def onClose(self):
        """"""
        self.destroy()
        self.original_frame.show()

if __name__ == "__main__":
    app = LoginApp()
    app.mainloop()
